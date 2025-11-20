import logging
import os
from typing import Any, List, Dict, Optional

from mcp.server.fastmcp import FastMCP

# Import exceptions
from excel_mcp.exceptions import (
    ValidationError,
    WorkbookError,
    SheetError,
    DataError,
    FormattingError,
    CalculationError,
    PivotError,
    ChartError
)

# Import from excel_mcp package with consistent _impl suffixes
from excel_mcp.validation import (
    validate_formula_in_cell_operation as validate_formula_impl,
    validate_range_in_sheet_operation as validate_range_impl
)
from excel_mcp.chart import create_chart_in_sheet as create_chart_impl
from excel_mcp.workbook import get_workbook_info
from excel_mcp.data import write_data
from excel_mcp.pivot import create_pivot_table as create_pivot_table_impl
from excel_mcp.tables import create_excel_table as create_table_impl
from excel_mcp.sheet import (
    copy_sheet,
    delete_sheet,
    rename_sheet,
    merge_range,
    unmerge_range,
    get_merged_ranges,
    insert_row,
    insert_cols,
    delete_rows,
    delete_cols,
)

# Get project root directory path for log file path.
# When using the stdio transmission method,
# relative paths may cause log files to fail to create
# due to the client's running location and permission issues,
# resulting in the program not being able to run.
# Thus using os.path.join(ROOT_DIR, "excel-mcp.log") instead.

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
LOG_FILE = os.path.join(ROOT_DIR, "excel-mcp.log")

# Initialize EXCEL_FILES_PATH variable without assigning a value
EXCEL_FILES_PATH = None

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        # Referring to https://github.com/modelcontextprotocol/python-sdk/issues/409#issuecomment-2816831318
        # The stdio mode server MUST NOT write anything to its stdout that is not a valid MCP message.
        logging.FileHandler(LOG_FILE)
    ],
)
logger = logging.getLogger("excel-mcp")
# Initialize FastMCP server
mcp = FastMCP(
    "excel-mcp",
    host=os.environ.get("FASTMCP_HOST", "0.0.0.0"),
    port=int(os.environ.get("FASTMCP_PORT", "8017")),
    instructions="Excel MCP Server for manipulating Excel files"
)

def get_excel_path(filename: str) -> str:
    """Get full path to Excel file.
    
    Args:
        filename: Name of Excel file
        
    Returns:
        Full path to Excel file
    """
    # If filename is already an absolute path, return it
    if os.path.isabs(filename):
        return filename

    # Check if in SSE mode (EXCEL_FILES_PATH is not None)
    if EXCEL_FILES_PATH is None:
        # Must use absolute path
        raise ValueError(f"Invalid filename: {filename}, must be an absolute path when not in SSE mode")

    # In SSE mode, if it's a relative path, resolve it based on EXCEL_FILES_PATH
    return os.path.join(EXCEL_FILES_PATH, filename)

@mcp.tool()
def apply_formula(
    filepath: str,
    sheet_name: str,
    cell: str,
    formula: str,
) -> str:
    """
    Apply an Excel formula to a cell with validation.
    Args:
      filepath (str): Path to the Excel file.
      sheet_name (str): Worksheet name.
      cell (str): Target cell address.
      formula (str): Excel formula to apply.
    Returns:
      message (str): Result of applying the formula or an error message.
    """
    try:
        full_path = get_excel_path(filepath)
        # First validate the formula
        validation = validate_formula_impl(full_path, sheet_name, cell, formula)
        if isinstance(validation, dict) and "error" in validation:
            return f"Error: {validation['error']}"
            
        # If valid, apply the formula
        from excel_mcp.calculations import apply_formula as apply_formula_impl
        result = apply_formula_impl(full_path, sheet_name, cell, formula)
        return result["message"]
    except (ValidationError, CalculationError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error applying formula: {e}")
        raise

@mcp.tool()
def validate_formula_syntax(
    filepath: str,
    sheet_name: str,
    cell: str,
    formula: str,
) -> str:
    """
    Validate Excel formula syntax without writing it to the sheet.
    Args:
      filepath (str): Path to the Excel file.
      sheet_name (str): Worksheet name.
      cell (str): Cell address for context.
      formula (str): Excel formula to validate.
    Returns:
      message (str): Validation result or an error message.
    """
    try:
        full_path = get_excel_path(filepath)
        result = validate_formula_impl(full_path, sheet_name, cell, formula)
        return result["message"]
    except (ValidationError, CalculationError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error validating formula: {e}")
        raise

@mcp.tool()
def format_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None,
    bold: bool = False,
    italic: bool = False,
    underline: bool = False,
    font_size: Optional[int] = None,
    font_color: Optional[str] = None,
    bg_color: Optional[str] = None,
    border_style: Optional[str] = None,
    border_color: Optional[str] = None,
    number_format: Optional[str] = None,
    alignment: Optional[str] = None,
    wrap_text: bool = False,
    merge_cells: bool = False,
    protection: Optional[Dict[str, Any]] = None,
    conditional_format: Optional[Dict[str, Any]] = None
) -> str:
    """
    Apply formatting options to a range of cells.

    Args:
      filepath (str): Path to the Excel file.
      sheet_name (str): Worksheet name.
      start_cell (str): Start cell of the range.
      end_cell (Optional[str]): End cell of the range.
      bold (bool): Whether to apply bold text.
      italic (bool): Whether to apply italic text.
      underline (bool): Whether to underline text.
      font_size (Optional[int]): Font size.
      font_color (Optional[str]): Font color in hex or name.
      bg_color (Optional[str]): Background color in hex or name.
      border_style (Optional[str]): Border style.
      border_color (Optional[str]): Border color.
      number_format (Optional[str]): Number format string.
      alignment (Optional[str]): Text alignment.
      wrap_text (bool): Whether to wrap text.
      merge_cells (bool): Whether to merge the range.
      protection (Optional[Dict[str, Any]]): Cell protection options.
      conditional_format (Optional[Dict[str, Any]]): Conditional format options.

    Returns:
      message (str): Result of the formatting operation or an error message.
    """
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.formatting import format_range as format_range_func
        
        # Convert None values to appropriate defaults for the underlying function
        format_range_func(
            filepath=full_path,
            sheet_name=sheet_name,
            start_cell=start_cell,
            end_cell=end_cell,  # This can be None
            bold=bold,
            italic=italic,
            underline=underline,
            font_size=font_size,  # This can be None
            font_color=font_color,  # This can be None
            bg_color=bg_color,  # This can be None
            border_style=border_style,  # This can be None
            border_color=border_color,  # This can be None
            number_format=number_format,  # This can be None
            alignment=alignment,  # This can be None
            wrap_text=wrap_text,
            merge_cells=merge_cells,
            protection=protection,  # This can be None
            conditional_format=conditional_format  # This can be None
        )
        return "Range formatted successfully"
    except (ValidationError, FormattingError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error formatting range: {e}")
        raise

@mcp.tool()
def read_data_from_excel(
    filepath: str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: Optional[str] = None,
    preview_only: bool = False
) -> str:
    """
    Read worksheet data with cell metadata and validation rules.

    Args:
      filepath (str): Path to the Excel file.
      sheet_name (str): Worksheet name.
      start_cell (str): Starting cell address.
      end_cell (Optional[str]): Ending cell address.
      preview_only (bool): Whether to return a preview subset.

    Returns:
      data (str): JSON string of cell data, positions, and validation info.
    """
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.data import read_excel_range_with_metadata
        result = read_excel_range_with_metadata(
            full_path, 
            sheet_name, 
            start_cell, 
            end_cell
        )
        if not result or not result.get("cells"):
            return "No data found in specified range"
            
        # Return as formatted JSON string
        import json
        return json.dumps(result, indent=2, default=str)
        
    except Exception as e:
        logger.error(f"Error reading data: {e}")
        raise

@mcp.tool()
def write_data_to_excel(
    filepath: str,
    sheet_name: str,
    data: List[List],
    start_cell: str = "A1",
) -> str:
    """
    Write tabular data to a worksheet starting at a given cell.

    Args:
      filepath (str): Path to the Excel file.
      sheet_name (str): Worksheet name.
      data (List[List]): Two-dimensional list of row data.
      start_cell (str): Top-left cell for writing.

    Returns:
      message (str): Result of the write operation or an error message.
    """
    try:
        full_path = get_excel_path(filepath)
        result = write_data(full_path, sheet_name, data, start_cell)
        return result["message"]
    except (ValidationError, DataError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error writing data: {e}")
        raise

@mcp.tool()
def create_workbook(filepath: str) -> str:
    """
    Create a new Excel workbook file.

    Args:
      filepath (str): Path where the workbook will be created.

    Returns:
      message (str): Creation result or an error message.
    """
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.workbook import create_workbook as create_workbook_impl
        create_workbook_impl(full_path)
        return f"Created workbook at {full_path}"
    except WorkbookError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating workbook: {e}")
        raise

@mcp.tool()
def create_worksheet(filepath: str, sheet_name: str) -> str:
    """
    Create a new worksheet in an existing workbook.

    Args:
      filepath (str): Path to the Excel file.
      sheet_name (str): Name of the new worksheet.

    Returns:
      message (str): Creation result or an error message.
    """
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.workbook import create_sheet as create_worksheet_impl
        result = create_worksheet_impl(full_path, sheet_name)
        return result["message"]
    except (ValidationError, WorkbookError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating worksheet: {e}")
        raise

@mcp.tool()
def create_chart(
    filepath: str,
    sheet_name: str,
    data_range: str,
    chart_type: str,
    target_cell: str,
    title: str = "",
    x_axis: str = "",
    y_axis: str = ""
) -> str:
    """
    Create a chart in a worksheet from a data range.

    Args:
      filepath (str): Path to the Excel file.
      sheet_name (str): Worksheet containing the data.
      data_range (str): Cell range used as chart data.
      chart_type (str): Chart type identifier.
      target_cell (str): Anchor cell for the chart.
      title (str): Chart title.
      x_axis (str): X-axis label.
      y_axis (str): Y-axis label.

    Returns:
      message (str): Chart creation result or an error message.
    """
    try:
        full_path = get_excel_path(filepath)
        result = create_chart_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            chart_type=chart_type,
            target_cell=target_cell,
            title=title,
            x_axis=x_axis,
            y_axis=y_axis
        )
        return result["message"]
    except (ValidationError, ChartError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating chart: {e}")
        raise

@mcp.tool()
def create_pivot_table(
    filepath: str,
    sheet_name: str,
    data_range: str,
    rows: List[str],
    values: List[str],
    columns: Optional[List[str]] = None,
    agg_func: str = "mean"
) -> str:
    """
    Create a pivot table in a worksheet from a data range.

    Args:
      filepath (str): Path to the Excel file.
      sheet_name (str): Worksheet containing the data.
      data_range (str): Source data range.
      rows (List[str]): Row field names.
      values (List[str]): Value field names.
      columns (Optional[List[str]]): Column field names.
      agg_func (str): Aggregation function name.

    Returns:
      message (str): Pivot table creation result or an error message.
    """
    try:
        full_path = get_excel_path(filepath)
        result = create_pivot_table_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            rows=rows,
            values=values,
            columns=columns or [],
            agg_func=agg_func
        )
        return result["message"]
    except (ValidationError, PivotError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating pivot table: {e}")
        raise

@mcp.tool()
def create_table(
    filepath: str,
    sheet_name: str,
    data_range: str,
    table_name: Optional[str] = None,
    table_style: str = "TableStyleMedium9"
) -> str:
    """
    Create a native Excel table from a cell range.

    Args:
      filepath (str): Path to the Excel file.
      sheet_name (str): Worksheet containing the range.
      data_range (str): Cell range to convert to a table.
      table_name (Optional[str]): Excel table name.
      table_style (str): Built-in table style name.

    Returns:
      message (str): Table creation result or an error message.
    """
    try:
        full_path = get_excel_path(filepath)
        result = create_table_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            table_name=table_name,
            table_style=table_style
        )
        return result["message"]
    except DataError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating table: {e}")
        raise

@mcp.tool()
def copy_worksheet(
    filepath: str,
    source_sheet: str,
    target_sheet: str
) -> str:
    """
    Copy a worksheet within the same workbook.

    Args:
      filepath (str): Path to the Excel file.
      source_sheet (str): Name of the worksheet to copy.
      target_sheet (str): Name for the new copied worksheet.

    Returns:
      message (str): Copy result or an error message.
    """
    try:
        full_path = get_excel_path(filepath)
        result = copy_sheet(full_path, source_sheet, target_sheet)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error copying worksheet: {e}")
        raise

@mcp.tool()
def delete_worksheet(
    filepath: str,
    sheet_name: str
) -> str:
    """
    Delete a worksheet from a workbook.

    Args:
      filepath (str): Path to the Excel file.
      sheet_name (str): Name of the worksheet to delete.

    Returns:
      message (str): Deletion result or an error message.
    """
    try:
        full_path = get_excel_path(filepath)
        result = delete_sheet(full_path, sheet_name)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting worksheet: {e}")
        raise

@mcp.tool()
def rename_worksheet(
    filepath: str,
    old_name: str,
    new_name: str
) -> str:
    """
    Rename an existing worksheet in a workbook.

    Args:
      filepath (str): Path to the Excel file.
      old_name (str): Current worksheet name.
      new_name (str): New worksheet name.

    Returns:
      message (str): Rename result or an error message.
    """
    try:
        full_path = get_excel_path(filepath)
        result = rename_sheet(full_path, old_name, new_name)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error renaming worksheet: {e}")
        raise

@mcp.tool()
def get_workbook_metadata(
    filepath: str,
    include_ranges: bool = False
) -> str:
    """
    Get workbook metadata such as sheets and named ranges.

    Args:
      filepath (str): Path to the Excel file.
      include_ranges (bool): Whether to include range information.

    Returns:
      info (str): Stringified metadata or an error message.
    """
    try:
        full_path = get_excel_path(filepath)
        result = get_workbook_info(full_path, include_ranges=include_ranges)
        return str(result)
    except WorkbookError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error getting workbook metadata: {e}")
        raise

@mcp.tool()
def merge_cells(filepath: str, sheet_name: str, start_cell: str, end_cell: str) -> str:
    """
    Merge a contiguous range of cells.

    Args:
      filepath (str): Path to the Excel file.
      sheet_name (str): Worksheet name.
      start_cell (str): Start cell address.
      end_cell (str): End cell address.

    Returns:
      message (str): Merge result or an error message.
    """
    try:
        full_path = get_excel_path(filepath)
        result = merge_range(full_path, sheet_name, start_cell, end_cell)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error merging cells: {e}")
        raise

@mcp.tool()
def unmerge_cells(filepath: str, sheet_name: str, start_cell: str, end_cell: str) -> str:
    """
    Unmerge a previously merged cell range.

    Args:
      filepath (str): Path to the Excel file.
      sheet_name (str): Worksheet name.
      start_cell (str): Start cell address.
      end_cell (str): End cell address.

    Returns:
      message (str): Unmerge result or an error message.
    """
    try:
        full_path = get_excel_path(filepath)
        result = unmerge_range(full_path, sheet_name, start_cell, end_cell)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error unmerging cells: {e}")
        raise

@mcp.tool()
def get_merged_cells(filepath: str, sheet_name: str) -> str:
    """
    List merged cell ranges in a worksheet.

    Args:
      filepath (str): Path to the Excel file.
      sheet_name (str): Worksheet name.

    Returns:
      info (str): Stringified merged ranges or an error message.
    """
    try:
        full_path = get_excel_path(filepath)
        return str(get_merged_ranges(full_path, sheet_name))
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error getting merged cells: {e}")
        raise

@mcp.tool()
def copy_range(
    filepath: str,
    sheet_name: str,
    source_start: str,
    source_end: str,
    target_start: str,
    target_sheet: Optional[str] = None
) -> str:
    """
    Copy a cell range to another position or sheet.

    Args:
      filepath (str): Path to the Excel file.
      sheet_name (str): Source worksheet name.
      source_start (str): Start cell of the source range.
      source_end (str): End cell of the source range.
      target_start (str): Top-left target cell.
      target_sheet (Optional[str]): Target worksheet name.

    Returns:
      message (str): Copy result or an error message.
    """
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.sheet import copy_range_operation
        result = copy_range_operation(
            full_path,
            sheet_name,
            source_start,
            source_end,
            target_start,
            target_sheet or sheet_name  # Use source sheet if target_sheet is None
        )
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error copying range: {e}")
        raise

@mcp.tool()
def delete_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str,
    shift_direction: str = "up"
) -> str:
    """
    Delete a cell range and shift remaining cells.

    Args:
      filepath (str): Path to the Excel file.
      sheet_name (str): Worksheet name.
      start_cell (str): Start cell of the range.
      end_cell (str): End cell of the range.
      shift_direction (str): Shift direction for remaining cells.

    Returns:
      message (str): Deletion result or an error message.
    """
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.sheet import delete_range_operation
        result = delete_range_operation(
            full_path,
            sheet_name,
            start_cell,
            end_cell,
            shift_direction
        )
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting range: {e}")
        raise

@mcp.tool()
def validate_excel_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None
) -> str:
    """
    Validate that a cell range exists and is well-formed.

    Args:
      filepath (str): Path to the Excel file.
      sheet_name (str): Worksheet name.
      start_cell (str): Start cell of the range.
      end_cell (Optional[str]): End cell of the range.

    Returns:
      message (str): Validation result or an error message.
    """
    try:
        full_path = get_excel_path(filepath)
        range_str = start_cell if not end_cell else f"{start_cell}:{end_cell}"
        result = validate_range_impl(full_path, sheet_name, range_str)
        return result["message"]
    except ValidationError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error validating range: {e}")
        raise

@mcp.tool()
def get_data_validation_info(
    filepath: str,
    sheet_name: str
) -> str:
    """
    Get all data validation rules defined in a worksheet.

    Args:
      filepath (str): Path to the Excel file.
      sheet_name (str): Worksheet name.

    Returns:
      rules (str): JSON string of validation rules or a message if none.
    """
    try:
        full_path = get_excel_path(filepath)
        from openpyxl import load_workbook
        from excel_mcp.cell_validation import get_all_validation_ranges
        
        wb = load_workbook(full_path, read_only=False)
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found"
            
        ws = wb[sheet_name]
        validations = get_all_validation_ranges(ws)
        wb.close()
        
        if not validations:
            return "No data validation rules found in this worksheet"
            
        import json
        return json.dumps({
            "sheet_name": sheet_name,
            "validation_rules": validations
        }, indent=2, default=str)
        
    except Exception as e:
        logger.error(f"Error getting validation info: {e}")
        raise

@mcp.tool()
def insert_rows(
    filepath: str,
    sheet_name: str,
    start_row: int,
    count: int = 1
) -> str:
    """
    Insert one or more rows at a given position.

    Args:
      filepath (str): Path to the Excel file.
      sheet_name (str): Worksheet name.
      start_row (int): First row index to insert at.
      count (int): Number of rows to insert.

    Returns:
      message (str): Insert result or an error message.
    """
    try:
        full_path = get_excel_path(filepath)
        result = insert_row(full_path, sheet_name, start_row, count)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error inserting rows: {e}")
        raise

@mcp.tool()
def insert_columns(
    filepath: str,
    sheet_name: str,
    start_col: int,
    count: int = 1
) -> str:
    """
    Insert one or more columns at a given position.

    Args:
      filepath (str): Path to the Excel file.
      sheet_name (str): Worksheet name.
      start_col (int): First column index to insert at.
      count (int): Number of columns to insert.

    Returns:
      message (str): Insert result or an error message.
    """
    try:
        full_path = get_excel_path(filepath)
        result = insert_cols(full_path, sheet_name, start_col, count)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error inserting columns: {e}")
        raise

@mcp.tool()
def delete_sheet_rows(
    filepath: str,
    sheet_name: str,
    start_row: int,
    count: int = 1
) -> str:
    """
    Delete one or more rows starting at a given position.

    Args:
      filepath (str): Path to the Excel file.
      sheet_name (str): Worksheet name.
      start_row (int): First row index to delete.
      count (int): Number of rows to delete.

    Returns:
      message (str): Deletion result or an error message.
    """
    try:
        full_path = get_excel_path(filepath)
        result = delete_rows(full_path, sheet_name, start_row, count)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting rows: {e}")
        raise

@mcp.tool()
def delete_sheet_columns(
    filepath: str,
    sheet_name: str,
    start_col: int,
    count: int = 1
) -> str:
    """
    Delete one or more columns starting at a given position.

    Args:
      filepath (str): Path to the Excel file.
      sheet_name (str): Worksheet name.
      start_col (int): First column index to delete.
      count (int): Number of columns to delete.

    Returns:
      message (str): Deletion result or an error message.
    """
    try:
        full_path = get_excel_path(filepath)
        result = delete_cols(full_path, sheet_name, start_col, count)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting columns: {e}")
        raise

def run_sse():
    """Run Excel MCP server in SSE mode."""
    # Assign value to EXCEL_FILES_PATH in SSE mode
    global EXCEL_FILES_PATH
    EXCEL_FILES_PATH = os.environ.get("EXCEL_FILES_PATH", "./excel_files")
    # Create directory if it doesn't exist
    os.makedirs(EXCEL_FILES_PATH, exist_ok=True)
    
    try:
        logger.info(f"Starting Excel MCP server with SSE transport (files directory: {EXCEL_FILES_PATH})")
        mcp.run(transport="sse")
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")

def run_streamable_http():
    """Run Excel MCP server in streamable HTTP mode."""
    # Assign value to EXCEL_FILES_PATH in streamable HTTP mode
    global EXCEL_FILES_PATH
    EXCEL_FILES_PATH = os.environ.get("EXCEL_FILES_PATH", "./excel_files")
    # Create directory if it doesn't exist
    os.makedirs(EXCEL_FILES_PATH, exist_ok=True)
    
    try:
        logger.info(f"Starting Excel MCP server with streamable HTTP transport (files directory: {EXCEL_FILES_PATH})")
        mcp.run(transport="streamable-http")
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")

def run_stdio():
    """Run Excel MCP server in stdio mode."""
    # No need to assign EXCEL_FILES_PATH in stdio mode
    
    try:
        logger.info("Starting Excel MCP server with stdio transport")
        mcp.run(transport="stdio")
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")